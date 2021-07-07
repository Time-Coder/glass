import os

try:
	import openpyxl as xl
	# from openpyxl.styles import Alignment
except ModuleNotFoundError as e:
	os.system("pip install openpyxl")
	import openpyxl as xl
	# from openpyxl.styles import Alignment

import datetime

def str2time(string):
	time_elements = string.split('/')
	return datetime.datetime(int(time_elements[0]), \
							 int(time_elements[1]), \
							 int(time_elements[2]), \
							 int(time_elements[3]), \
							 int(time_elements[4]), \
							 int(time_elements[5]))

class Excel:
	def __init__(self, filename = None, sheetname = 0):
		self.filename = filename
		self.current_sheet = None

		if filename is None or not os.path.isfile(self.filename):
			self.is_new = True
			self.workbook = xl.Workbook()	
		else:
			self.is_new = False
			self.workbook = xl.load_workbook(self.filename)
			
		self.sheet(sheetname)

	def sheet(self, sheetname):
		if isinstance(sheetname, str):
			if sheetname in self.workbook.sheetnames:
				self.current_sheet = self.workbook[sheetname]
			else:
				self.current_sheet = self.workbook.create_sheet(sheetname)
		elif isinstance(sheetname, int) and sheetname >= 0:
			if sheetname < len(self.workbook.sheetnames):
				self.current_sheet = self.workbook.worksheets[sheetname]
			else:
				for i in range(len(self.workbook.sheetnames) + 1, sheetname):
					self.workbook.create_sheet("Sheet" + str(i))
				self.current_sheet = self.workbook.worksheets[sheetname]

	def rows(self):
		if self.current_sheet is None:
			return 0
		else:
			return self.current_sheet.max_row

	def cols(self):
		if self.current_sheet is None:
			return 0
		else:
			return self.current_sheet.max_column

	def min_row(self):
		if self.current_sheet is None:
			return -1
		else:
			return self.current_sheet.min_row-1

	def max_row(self):
		if self.current_sheet is None:
			return -1
		else:
			return self.current_sheet.max_row-1

	def min_col(self):
		if self.current_sheet is None:
			return -1
		else:
			return self.current_sheet.min_column-1

	def max_col(self):
		if self.current_sheet is None:
			return -1
		else:
			return self.current_sheet.max_column-1

	def valid_rows(self):
		if self.current_sheet is None:
			return 0
		else:
			return self.current_sheet.max_row - self.current_sheet.min_row + 1

	def valid_column(self):
		if self.current_sheet is None:
			return 0
		else:
			return self.current_sheet.max_column - self.current_sheet.min_column + 1

	def __getitem__(self, position):
		i, j = position
		return self.current_sheet.cell(row=i+1, column=j+1).value

	def __setitem__(self, position, value):
		i, j = position

		if isinstance(value, str) and '\n' in value:
			self.current_sheet.cell(row=i+1, column=j+1).alignment = xl.styles.Alignment(wrapText=True)

		self.current_sheet.cell(row=i+1, column=j+1).value = value

	def save(self, filename = None):
		if filename is None:
			filename = self.filename

		if filename is None:
			return

		not_close = True
		info_printed = False
		while not_close:
			try:
				self.workbook.save(filename)
				not_close = False
			except Exception as e:
				if not info_printed:
					print("Please close " + filename + " now! I'm waiting...")
					info_printed = True

	def merge(self, start_row, end_row, start_col, end_col):
		self.current_sheet.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=end_row+1, end_column=end_col+1)